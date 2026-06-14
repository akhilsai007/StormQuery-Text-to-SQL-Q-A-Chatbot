"""
exporters.py
Turn a chatbot query result into downloadable CSV, Excel, and PDF files.

Each function takes the query-result DataFrame plus optional metadata (the
user's question, the SQL that was run, the natural-language answer) and returns
the file as raw bytes -- exactly what Streamlit's st.download_button wants.

Design notes:
- Everything is built in memory (io.BytesIO). No temp files to clean up.
- The heavy imports (openpyxl, reportlab) live inside their functions, so a
  CSV-only user doesn't need them installed.

Dependencies:
    pip install pandas openpyxl reportlab
"""

from __future__ import annotations

import io
from datetime import datetime
from xml.sax.saxutils import escape as _xml_escape

import pandas as pd


def _esc(text) -> str:
    """Escape &, <, > so reportlab Paragraphs don't mis-parse them as markup."""
    return _xml_escape(str(text))


# ----------------------------------------------------------------------
# CSV  -- raw, full-fidelity data
# ----------------------------------------------------------------------
def to_csv_bytes(df: pd.DataFrame) -> bytes:
    """Return the DataFrame as UTF-8 CSV bytes."""
    return df.to_csv(index=False).encode("utf-8")


# ----------------------------------------------------------------------
# Excel (.xlsx)  -- styled data + an optional provenance sheet
# ----------------------------------------------------------------------
def to_excel_bytes(
    df: pd.DataFrame,
    question: str | None = None,
    sheet_name: str = "Results",
) -> bytes:
    """
    Return the DataFrame as a styled .xlsx workbook.

    The data sits on the 'Results' sheet with a bold, colored, frozen header
    row and auto-sized columns. If a question is given, a 'Query info' sheet
    records it for context.
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

        # Auto-size each column to its longest value (capped so it stays sane).
        for col in ws.columns:
            longest = max(
                (len(str(c.value)) for c in col if c.value is not None),
                default=10,
            )
            ws.column_dimensions[col[0].column_letter].width = min(longest + 2, 60)

        ws.freeze_panes = "A2"  # keep the header visible while scrolling

        if question:
            info = writer.book.create_sheet("Query info")
            rows = [
                ("Question", question or ""),
                ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ]
            for r, (label, value) in enumerate(rows, start=1):
                info[f"A{r}"] = label
                info[f"B{r}"] = value
                info[f"A{r}"].font = Font(bold=True)
            info.column_dimensions["A"].width = 14
            info.column_dimensions["B"].width = 90

    return buffer.getvalue()


# ----------------------------------------------------------------------
# PDF  -- a readable report (title, question, answer, table, SQL)
# ----------------------------------------------------------------------
def to_pdf_bytes(
    df: pd.DataFrame,
    question: str | None = None,
    answer: str | None = None,
    chart_png: bytes | None = None,
    title: str = "Storm Events Query Result",
    max_rows: int = 100,
    max_cell_chars: int = 45,
) -> bytes:
    """
    Return a formatted PDF report as bytes: title, question, answer, an optional
    chart image (chart_png), then the data table.

    A PDF is a *readable* artifact, not a data dump: long tables are capped at
    `max_rows` and long cell values are truncated for display. Point users to
    the CSV/Excel export for full fidelity.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import (
        Image as RLImage,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),  # landscape gives wide tables room
        topMargin=0.6 * inch,
        bottomMargin=0.6 * inch,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("Label", parent=styles["Normal"],
                              fontName="Helvetica-Bold", spaceAfter=4))

    story = [Paragraph(_esc(title), styles["Title"]), Spacer(1, 8)]

    if question:
        story += [Paragraph("Question", styles["Label"]),
                  Paragraph(_esc(question), styles["Normal"]), Spacer(1, 8)]
    if answer:
        story += [Paragraph("Answer", styles["Label"]),
                  Paragraph(_esc(answer), styles["Normal"]), Spacer(1, 12)]

    # Embed the visualization, if one was supplied. Sized to preserve the PNG's
    # aspect ratio. Wrapped so a bad image can never break the report.
    if chart_png:
        try:
            iw, ih = ImageReader(io.BytesIO(chart_png)).getSize()
            display_w = 5.2 * inch
            img = RLImage(io.BytesIO(chart_png), width=display_w,
                          height=display_w * ih / iw)
            img.hAlign = "CENTER"
            story += [img, Spacer(1, 12)]
        except Exception:
            pass

    # Build the table, truncating rows and long cell values for readability.
    shown = df.head(max_rows)

    def _clip(v) -> str:
        # Coerce every cell to text: NULL/NaN -> blank, everything else -> str.
        s = "" if pd.isna(v) else str(v)
        return s if len(s) <= max_cell_chars else s[: max_cell_chars - 1] + "\u2026"

    data = [list(shown.columns)] + [[_clip(v) for v in row] for row in shown.values]

    table = Table(data, repeatRows=1)  # repeat the header on every page
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F2F2F2")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#BFBFBF")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(table)

    if len(df) > max_rows:
        story += [Spacer(1, 6), Paragraph(
            f"Showing first {max_rows:,} of {len(df):,} rows. "
            "Download the CSV or Excel file for the full result.",
            styles["Italic"])]

    doc.build(story)
    return buffer.getvalue()
